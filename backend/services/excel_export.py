"""
services/excel_export.py
Export attendance records to a premium multi-sheet Excel file.
"""

import io
from datetime import date
from typing import Optional

import pandas as pd
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from backend.models.attendance import AttendanceRecord
from backend.models.student import Student


async def build_attendance_excel(
    db: AsyncSession,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
) -> bytes:
    query = (
        select(
            AttendanceRecord.roll_no,
            Student.name,
            AttendanceRecord.date,
            AttendanceRecord.status,
            AttendanceRecord.confidence,
        )
        .join(Student, Student.roll_no == AttendanceRecord.roll_no, isouter=True)
        .order_by(AttendanceRecord.date, AttendanceRecord.roll_no)
    )

    if start_date:
        query = query.where(AttendanceRecord.date >= start_date)
    if end_date:
        query = query.where(AttendanceRecord.date <= end_date)

    result = await db.execute(query)
    rows = result.all()

    df = pd.DataFrame(
        rows,
        columns=["roll_no", "name", "date", "status", "confidence"],
    )

    buffer = io.BytesIO()

    if df.empty:
        # Create empty excel
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            pd.DataFrame(columns=["roll_no", "name", "date", "status"]).to_excel(
                writer, index=False, sheet_name="Attendance"
            )
        buffer.seek(0)
        return buffer.getvalue()

    # Normalize date as string YYYY-MM-DD
    df["date"] = df["date"].astype(str)

    # 2. Pivot for Attendance Sheet
    df_pivot = df.pivot(index=["roll_no", "name"], columns="date", values="status").reset_index()
    # Replace NaNs with 'A' (Absent)
    df_pivot.fillna("A", inplace=True)

    # Rename columns for presentation
    df_pivot.rename(columns={"roll_no": "Roll No", "name": "Name"}, inplace=True)

    # 3. Create Summary Sheet
    summary_data = []
    grouped = df.groupby(["roll_no", "name"])
    for (roll_no, name), group in grouped:
        total = len(group)
        present = sum(group["status"] == "P")
        absent = sum(group["status"] == "A")
        pct = round((present / total) * 100, 1) if total > 0 else 0.0
        summary_data.append({
            "Roll No": roll_no,
            "Name": name,
            "Total Days": total,
            "Present Days": present,
            "Absent Days": absent,
            "Attendance %": pct,
        })
    df_summary = pd.DataFrame(summary_data)
    df_summary.sort_values(by="Roll No", inplace=True)

    # Write to excel using ExcelWriter and openpyxl styling
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_pivot.to_excel(writer, index=False, sheet_name="Attendance")
        df_summary.to_excel(writer, index=False, sheet_name="Summary")

        workbook = writer.book

        # Styles
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")  # Dark slate
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")

        present_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")  # Light green
        present_font = Font(name="Segoe UI", size=10, color="15803D", bold=True)

        absent_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")  # Light red
        absent_font = Font(name="Segoe UI", size=10, color="B91C1C", bold=True)

        regular_font = Font(name="Segoe UI", size=10)
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")

        thin_border = Border(
            left=Side(style="thin", color="E5E7EB"),
            right=Side(style="thin", color="E5E7EB"),
            top=Side(style="thin", color="E5E7EB"),
            bottom=Side(style="thin", color="E5E7EB"),
        )

        # Style Attendance Sheet
        ws_att = workbook["Attendance"]
        ws_att.views.sheetView[0].showGridLines = True

        # Header formatting
        for col_idx in range(1, ws_att.max_column + 1):
            cell = ws_att.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        # Data formatting
        for r_idx in range(2, ws_att.max_row + 1):
            for c_idx in range(1, ws_att.max_column + 1):
                cell = ws_att.cell(row=r_idx, column=c_idx)
                cell.font = regular_font
                cell.border = thin_border

                # Check status columns (c_idx >= 3)
                if c_idx >= 3:
                    cell.alignment = center_align
                    val = str(cell.value)
                    if val == "P":
                        cell.value = "Present"
                        cell.fill = present_fill
                        cell.font = present_font
                    elif val == "A":
                        cell.value = "Absent"
                        cell.fill = absent_fill
                        cell.font = absent_font
                else:
                    cell.alignment = left_align

        # Auto-adjust column widths
        for col in ws_att.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_att.column_dimensions[col_letter].width = max(max_len + 4, 12)

        # Style Summary Sheet
        ws_sum = workbook["Summary"]
        ws_sum.views.sheetView[0].showGridLines = True

        # Header formatting
        for col_idx in range(1, ws_sum.max_column + 1):
            cell = ws_sum.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        # Data formatting
        for r_idx in range(2, ws_sum.max_row + 1):
            for c_idx in range(1, ws_sum.max_column + 1):
                cell = ws_sum.cell(row=r_idx, column=c_idx)
                cell.font = regular_font
                cell.border = thin_border

                # Format percentage column (c_idx = 6)
                if c_idx == 6:
                    val = cell.value
                    if isinstance(val, (int, float)):
                        cell.value = f"{val}%"
                    cell.alignment = center_align
                elif c_idx in [3, 4, 5]:
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align

        # Auto-adjust column widths
        for col in ws_sum.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_sum.column_dimensions[col_letter].width = max(max_len + 4, 14)

    buffer.seek(0)
    return buffer.getvalue()
